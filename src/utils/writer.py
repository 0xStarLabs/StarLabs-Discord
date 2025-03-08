import asyncio
from openpyxl import load_workbook
from loguru import logger
from src.utils.constants import ACCOUNTS_FILE

# Создаем глобальный асинхронный lock
file_lock = asyncio.Lock()


async def update_account(token: str, field: str, value: str) -> bool:
    """
    Безопасно обновляет поле аккаунта в XLSX файле.

    Args:
        token (str): Discord токен аккаунта
        field (str): Название поля для обновления (DISCORD_TOKEN, PROXY, USERNAME, STATUS)
        value (str): Новое значение поля

    Returns:
        bool: True если обновление успешно, False если аккаунт не найден
    """
    # Маппинг названий полей к индексам колонок
    field_mapping = {
        "DISCORD_TOKEN": 0,
        "PROXY": 1, 
        "USERNAME": 2, 
        "STATUS": 3, 
        "PASSWORD": 4, 
        "NEW_PASSWORD": 5, 
        "NEW_NAME": 6, 
        "NEW_USERNAME": 7, 
        "NEW_PROFILE_PICTURE": 8
    }

    if field not in field_mapping:
        logger.error(f"Invalid field name: {field}")
        return False

    column_index = field_mapping[field]

    async with file_lock:  # Используем асинхронный контекстный менеджер
        try:
            # Используем loop.run_in_executor для блокирующих операций с файлом
            loop = asyncio.get_event_loop()
            workbook = await loop.run_in_executor(
                None, load_workbook, ACCOUNTS_FILE
            )
            sheet = workbook.active

            # Ищем строку с нужным токеном
            token_found = False
            for row_index, row in enumerate(sheet.rows, 1):
                if row[0].value == token:
                    # Обновляем значение в нужной ячейке
                    sheet.cell(row=row_index, column=column_index + 1, value=value)
                    token_found = True
                    break

            if not token_found:
                logger.error(f"Account with token {token[:10]}... not found")
                await loop.run_in_executor(None, workbook.close)
                return False

            # Сохраняем изменения
            await loop.run_in_executor(None, workbook.save, ACCOUNTS_FILE)
            await loop.run_in_executor(None, workbook.close)

            logger.success(f"Successfully updated {field} for account {token[:10]}...")
            return True

        except Exception as e:
            logger.error(f"Error updating account: {str(e)}")
            return False
